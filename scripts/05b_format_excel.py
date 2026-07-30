"""
05b_format_excel.py
Apply professional formatting to the summary workbook: Arial font, bold
headers, currency/percent number formats, autosized columns.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PATH = "outputs/Stock_Market_Summary_Tables.xlsx"
wb = openpyxl.load_workbook(PATH)

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)

PCT_COLS = {"Annualized_Return", "Annualized_Volatility", "Total_Return",
            "Avg_Annualized_Return", "Avg_Annualized_Volatility", "Avg_Daily_Return"}
CURRENCY_COLS = {"First_Close", "Last_Close", "Avg_Close"}
INT_COLS = {"Trading_Days", "Num_Tickers", "Active_Tickers", "Total_Volume", "Avg_Daily_Volume"}

for ws in wb.worksheets:
    headers = [c.value for c in ws[1]]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

        col_letter = get_column_letter(col_idx)
        max_len = len(str(header)) + 2
        for row_idx in range(2, ws.max_row + 1):
            body_cell = ws.cell(row=row_idx, column=col_idx)
            body_cell.font = BODY_FONT
            if header in PCT_COLS and isinstance(body_cell.value, (int, float)):
                body_cell.number_format = "0.0%"
            elif header in CURRENCY_COLS and isinstance(body_cell.value, (int, float)):
                body_cell.number_format = "$#,##0.00"
            elif header in INT_COLS and isinstance(body_cell.value, (int, float)):
                body_cell.number_format = "#,##0"
            val_len = len(str(body_cell.value)) if body_cell.value is not None else 0
            max_len = max(max_len, val_len)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 22)
    ws.freeze_panes = "A2"

wb.save(PATH)
print("Formatted workbook saved:", PATH)
